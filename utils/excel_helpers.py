"""
amber/utils/excel_helpers.py
────────────────────────────
Helpers Excel para download de template e leitura de dados.
Idêntico à Célula 4 da versão Spyder.
Requer: pip install openpyxl
"""
from __future__ import annotations

import io


def generate_template(t_mission_ref: float = 27.0) -> bytes:
    """
    Gera o arquivo AMBER_template.xlsx em memória e retorna os bytes.
    Usado na view de download para enviar o arquivo ao navegador.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("Instale openpyxl:  pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observacoes"

    # ── Cabeçalho ──────────────────────────────────────────────────────────
    headers = [
        "Equipamento",
        f"t_obs (anos, máx={t_mission_ref:.1f})",
        "Falha (1=sim, 0=nao)",
    ]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="185FA5")
    hdr_font = Font(bold=True, color="FFFFFF")
    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    for col in range(1, 4):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center')
        c.border = borda

    # ── Dados de exemplo ───────────────────────────────────────────────────
    ambar_fill = PatternFill("solid", fgColor="FAEEDA")
    for linha in [
        (1,  5.0,            0),
        (2,  8.5,            1),   # falha
        (3, 10.0,            0),
        (4,  3.2,            1),   # falha
        (5, t_mission_ref,   0),
    ]:
        ws.append(list(linha))
        for col in range(1, 4):
            c = ws.cell(row=ws.max_row, column=col)
            if col == 3 and linha[2] == 1:
                c.fill = PatternFill("solid", fgColor="FFDAD6")  # vermelho falha
            else:
                c.fill = ambar_fill
            c.border = borda
            c.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 22

    # ── Aba de instruções ──────────────────────────────────────────────────
    wi = wb.create_sheet("Instrucoes")
    for row in [
        ["AMBER v2.1 — Instruções de preenchimento"],
        [""],
        ["Coluna A:", "Índice do equipamento (ignorado na leitura)."],
        ["Coluna B:", "Tempo de observação t_obs em anos. Deve ser ≤ T_MISSION."],
        ["Coluna C:", "0 = censurado (sobreviveu a t_obs)  |  1 = falhou em t_obs"],
        [""],
        ["Modelo:",    "k_i = (t_obs_i / T_MISSION)^β_W"],
        ["Censurado:", "contribui k_i para α  →  evidência A FAVOR de R alto"],
        ["Falha:",     "contribui k_i para β  →  evidência CONTRA R alto"],
        ["Posterior:", "Beta( α₀ + Σk_i[censurados],  β₀ + Σk_j[falhas] )"],
    ]:
        wi.append(row)
    wi.column_dimensions['A'].width = 14
    wi.column_dimensions['B'].width = 65

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_excel(file_bytes: bytes) -> list[tuple[float, bool]]:
    """
    Lê os bytes de um arquivo Excel com o formato AMBER e retorna
    lista de (t_obs_i, is_failure). Ignora cabeçalho e linhas vazias.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Instale openpyxl:  pip install openpyxl")

    buf = io.BytesIO(file_bytes)
    wb  = openpyxl.load_workbook(buf, data_only=True)
    ws  = wb.active

    obs: list[tuple[float, bool]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        t_i    = float(row[1])
        fail_i = bool(int(row[2])) if row[2] is not None else False
        obs.append((t_i, fail_i))
    return obs
